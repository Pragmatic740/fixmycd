"""
Convert US Bridge Inventory Excel → validated JSON for seed-nbi-bridges.ts.

Usage:
  python scripts/nbi_xlsx_to_json.py
  python scripts/nbi_xlsx_to_json.py --input "context_file/US_Bridge_Inventory_CD_2026V14 (1).xlsx"
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "context_file" / "US_Bridge_Inventory_CD_2026V14 (1).xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "nbi-bridges-2026v14.json"

LINK_RE = re.compile(r'url="([^"]+)"', re.IGNORECASE)


def parse_maps_url(h_link: object) -> str | None:
    if not h_link or not isinstance(h_link, str):
        return None
    m = LINK_RE.search(h_link)
    if m:
        return m.group(1)
    if h_link.startswith("http"):
        return h_link
    return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert NBI bridge xlsx to JSON")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"Input not found: {args.input}")

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    if "NBI_CD3" not in wb.sheetnames:
        raise SystemExit(f"Expected sheet NBI_CD3, found {wb.sheetnames}")

    ws = wb["NBI_CD3"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}

    required = [
        "NBI ID",
        "LATDD",
        "LONGDD",
        "Condition",
        "Repair Cost",
        "Con. Dist.",
        "Road Carried",
        "Features",
        "Location",
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        raise SystemExit(f"Missing columns: {missing}")

    bridges = []
    skipped = 0
    conditions: Counter[str] = Counter()
    districts: Counter[str] = Counter()
    null_coords = 0
    costs: list[float] = []

    for row in rows:
        nbi_id = row[idx["NBI ID"]]
        lat = row[idx["LATDD"]]
        lng = row[idx["LONGDD"]]
        if nbi_id is None or lat is None or lng is None:
            skipped += 1
            if lat is None or lng is None:
                null_coords += 1
            continue

        condition = str(row[idx["Condition"]] or "").strip().upper() or "P"
        cost_raw = row[idx["Repair Cost"]]
        repair_cost = float(cost_raw) if isinstance(cost_raw, (int, float)) else None
        cd = str(row[idx["Con. Dist."]] or "").strip() or None
        year_built = row[idx["Year Built"]] if "Year Built" in idx else None
        last_work = row[idx["Last Work"]] if "Last Work" in idx else None

        record = {
            "objectId": row[idx["OBJECTID"]] if "OBJECTID" in idx else None,
            "nbiId": str(nbi_id).strip(),
            "congressionalDistrict": cd,
            "routeNo": str(row[idx["Route No."]]).strip() if "Route No." in idx and row[idx["Route No."]] is not None else None,
            "features": str(row[idx["Features"]] or "").strip() or None,
            "roadCarried": str(row[idx["Road Carried"]] or "").strip() or None,
            "location": str(row[idx["Location"]] or "").strip() or None,
            "yearBuilt": int(year_built) if isinstance(year_built, (int, float)) and year_built else None,
            "fc": str(row[idx["FC"]]).strip() if "FC" in idx and row[idx["FC"]] is not None else None,
            "lastWork": int(last_work) if isinstance(last_work, (int, float)) and last_work else 0,
            "condition": condition,
            "latitude": float(lat),
            "longitude": float(lng),
            "cdRank": row[idx["CD Rank"]] if "CD Rank" in idx else None,
            "repairCost": repair_cost,
            "mapsUrl": parse_maps_url(row[idx["H_Link"]]) if "H_Link" in idx else None,
        }
        bridges.append(record)
        conditions[condition] += 1
        if cd:
            districts[cd] += 1
        if repair_cost is not None:
            costs.append(repair_cost)

    payload = {
        "datasetKey": "nbi-bridges-2026v14",
        "source": args.input.name,
        "generatedFromSheet": "NBI_CD3",
        "count": len(bridges),
        "bridges": bridges,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(f"Wrote {len(bridges)} bridges -> {args.output}")
    print(f"Skipped: {skipped} (null coords: {null_coords})")
    print(f"Condition: {dict(conditions)}")
    print(f"Congressional districts: {len(districts)}")
    print(f"Top CDs: {districts.most_common(8)}")
    if costs:
        print(
            f"Repair cost min/avg/max: "
            f"${min(costs):,.0f} / ${sum(costs)/len(costs):,.0f} / ${max(costs):,.0f}"
        )


if __name__ == "__main__":
    main()
